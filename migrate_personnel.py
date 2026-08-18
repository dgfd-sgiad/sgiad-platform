# -*- coding: utf-8 -*-
"""Migration complete PERSONNEL.xlsx -> conges_agents (sans notes)"""
import openpyxl
from datetime import datetime
from db import get_supabase

wb = openpyxl.load_workbook('PERSONNEL.xlsx', data_only=True)
ws = wb['BASE EN COURS'] if 'BASE EN COURS' in wb.sheetnames else wb.active

rows = list(ws.iter_rows(values_only=True))
hi = next(i for i, r in enumerate(rows[:10]) if any('matricule' in str(c).lower() for c in r if c))
headers = [str(c).strip().lower() if c is not None else '' for c in rows[hi]]

def idx(*keys):
    for i, h in enumerate(headers):
        if all(k in h for k in keys):
            return i
    return None

I_MAT   = idx('matricule')
I_NOM   = idx('nom')
I_NPI   = idx('ifu')
I_STR   = idx('structure')
I_SEXE  = idx('sexe')
I_POSTE = idx('poste')
I_STAT  = idx('statut')
I_RETR  = idx('probable')
I_ANRET = idx('année de départ')
I_NAIS  = idx('naissance')
I_PSFP  = idx('prise de service dans la fonction')
I_PSST  = idx('structure actuelle')
I_ANC   = idx('ancienneté')
I_CORPS = idx('corps')
I_GRADE = idx('grade actuelle')
I_GPAYE = idx('payé')
I_CAT   = idx('catégorie')
I_CONT  = idx('contact')
I_DIP   = idx('dernier diplôme')
I_DIPR  = idx('reconnu')
I_FORM  = idx('formation diplômante')
I_DSI   = idx('interrompu')
I_PER   = idx('periode')
I_NBJ   = idx('nombre de jour')
I_OBS   = idx('observation')

COLS = [
    ('nom', I_NOM), ('npi_ifu', I_NPI), ('direction', I_STR), ('sexe', I_SEXE),
    ('poste', I_POSTE), ('statut_admin', I_STAT), ('date_retraite', I_RETR),
    ('annee_retraite', I_ANRET), ('date_naissance', I_NAIS), ('date_prise_service', I_PSFP),
    ('prise_service_structure', I_PSST), ('anciennete_fp', I_ANC), ('corps', I_CORPS),
    ('grade', I_GRADE), ('grade_paye', I_GPAYE), ('cat_admin', I_CAT), ('contact', I_CONT),
    ('diplome', I_DIP), ('diplome_reconnu', I_DIPR), ('date_formation', I_FORM),
    ('service_interrompu', I_DSI), ('periode_dernier_conge', I_PER),
    ('nb_jours_dernier_conge', I_NBJ), ('observation', I_OBS),
]

def fmt(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    return s if s else None

sb = get_supabase()
maj, ins = 0, 0

for r in rows[hi+1:]:
    mat = fmt(r[I_MAT])
    if not mat or not mat.isdigit():
        continue
    payload = {}
    for db_col, i in COLS:
        if i is not None:
            v = fmt(r[i])
            if v is not None:
                payload[db_col] = v

    existing = sb.table('conges_agents').select('id').eq('matricule', mat).execute().data
    if existing:
        sb.table('conges_agents').update(payload).eq('matricule', mat).execute()
        maj += 1
    else:
        nom = payload.get('nom', '')
        last2 = ' '.join(nom.split(' ')[:2])
        dup = sb.table('conges_agents').select('id').ilike('nom', last2 + '%').execute().data if last2 else []
        if len(dup) == 1:
            payload['matricule'] = mat
            sb.table('conges_agents').update(payload).eq('id', dup[0]['id']).execute()
            maj += 1
        else:
            payload['matricule'] = mat
            payload['statut'] = 'actif'
            payload['categorie'] = 'fonctionnaire'
            new = sb.table('conges_agents').insert(payload).execute().data[0]
            sb.table('conges_droits').insert({
                'agent_id': new['id'], 'annee': 2026,
                'droit_annuel': 30, 'report': 0, 'disponible': 30,
                'consomme': 0, 'solde': 30,
            }).execute()
            ins += 1

print(f'✔ Migration terminée : {maj} mises à jour, {ins} nouveaux agents')
