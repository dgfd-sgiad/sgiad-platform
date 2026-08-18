# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('PERSONNEL.xlsx', data_only=True)
print(f"Feuilles disponibles : {wb.sheetnames}")
ws = wb['BASE EN COURS'] if 'BASE EN COURS' in wb.sheetnames else wb.active
print(f"Feuille active : {ws.title}")

rows = list(ws.iter_rows(values_only=True))
print(f"Nombre total de lignes : {len(rows)}")

print("\n=== 10 premières lignes ===")
for i, r in enumerate(rows[:10]):
    print(f"Ligne {i}: {r[:5]}")

# Cherche la ligne d'en-tête
hi = None
for i, r in enumerate(rows[:10]):
    if any('matricule' in str(c).lower() for c in r if c):
        hi = i
        break

print(f"\nLigne d'en-tête détectée : {hi}")
if hi is not None:
    print(f"En-tête complet : {rows[hi]}")
    print(f"\nLigne suivante (données) : {rows[hi+1][:5]}")
    