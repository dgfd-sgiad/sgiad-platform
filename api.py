# -*- coding: utf-8 -*-
"""
SGIAD - API Flask (v2 Supabase)
================================
Serveur principal. Les donnees principales (accords, localisation,
parametres, suivi) viennent de Supabase (PostgreSQL cloud).
Les modules secondaires (coop decentralisee, anciens accords) restent
sur Excel jusqu'a leur migration.

Variables d'environnement:
    SUPABASE_URL  = https://xxxxx.supabase.co
    SUPABASE_KEY  = eyJhbGciOi...
"""

import csv
import re
import traceback
import unicodedata
import numpy as np
from flask import Flask, request, jsonify, send_from_directory, send_file, make_response
from flask_cors import CORS
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from io import BytesIO, StringIO
from datetime import datetime
from urllib.parse import unquote
import pandas as pd
import json
import os
import sys
import time
import threading
import tempfile

# Charge le fichier .env AVANT que auth.py lise SIGNUP_ENABLED
from dotenv import load_dotenv
load_dotenv()


app = Flask(__name__)

# CORS restreint aux domaines autorises (le frontend est servi en same-origin)
_default_origins = [
    'https://sgiad-platform.onrender.com',
    'http://localhost:5000',
    'http://127.0.0.1:5000',
]
_cors_origins = [o.strip() for o in os.environ.get('CORS_ORIGINS', '').split(',') if o.strip()] or _default_origins
CORS(app, resources={r"/api/*": {"origins": _cors_origins}})

# ── Module Gestion des Accords Financiers (Supabase) ──
from accords_financiers import bp as accords_financiers_bp
app.register_blueprint(accords_financiers_bp)

# ── Module Authentification ──
from auth import bp as auth_bp, require_auth
app.register_blueprint(auth_bp)


@app.after_request
def add_security_headers(resp):
    """En-tetes de securite sur toutes les reponses."""
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    resp.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    resp.headers.setdefault('Content-Security-Policy', "default-src 'self'; script-src 'self' 'unsafe-inline' https://unpkg.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://unpkg.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; font-src 'self' https://fonts.gstatic.com https://unpkg.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; img-src 'self' data: https://images.unsplash.com https://tile.openstreetmap.org https://*.tile.openstreetmap.org https://unpkg.com https://cdn.jsdelivr.net; connect-src 'self' https://unpkg.com https://cdnjs.cloudflare.com https://tile.openstreetmap.org https://*.tile.openstreetmap.org;")
    if request.path.startswith('/api/'):
        resp.headers.setdefault('Cache-Control', 'no-store')
    return resp

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

BANQUE_FILE   = os.path.join(BASE_DIR, 'Banque_Projets.xlsx')
ACCORDS_FILE  = os.path.join(BASE_DIR, 'Banque_Accords_V2.xlsx')

# ── Servir les fichiers statiques (HTML, CSS, JS, assets) ──
@app.route('/')
def accueil_public():
    """Page d'accueil publique de la plateforme (visible par tous)."""
    resp = send_file(os.path.join(BASE_DIR, 'accueil.html'))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return resp

@app.route('/app')
def app_interne():
    """Application DGFD accessible apres connexion depuis l'accueil."""
    resp = send_file(os.path.join(BASE_DIR, 'index.html'))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return resp


@app.route('/open-data')
def open_data_page():
    """Page publique d'exploration Open Data filtrée."""
    return send_file(os.path.join(BASE_DIR, 'open_data.html'))


# Extensions autorisees pour les fichiers statiques (pas de code source,
# pas de secrets .env, pas de donnees .xlsx/.json)
ALLOWED_STATIC_EXT = {'.html', '.css', '.js', '.png', '.jpg', '.jpeg', '.webp',
                      '.svg', '.ico', '.gif', '.geojson', '.pdf', '.woff', '.woff2'}
BLOCKED_STATIC_NAMES = {'session_log.json'}

@app.route('/<path:filepath>')
def serve_static(filepath):
    """Sert uniquement les fichiers statiques autorises (HTML, CSS, JS, images).
    Bloque le code source (.py), les secrets (.env) et les donnees (.xlsx)."""
    filename = os.path.basename(filepath)
    if filename.startswith('.') or filename in BLOCKED_STATIC_NAMES:
        return jsonify({"error": "Fichier introuvable"}), 404
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ALLOWED_STATIC_EXT:
        return jsonify({"error": "Fichier introuvable"}), 404
    # Protege contre toute tentative de remontee d'arborescence
    full_path = os.path.realpath(os.path.join(BASE_DIR, filepath))
    if not full_path.startswith(os.path.realpath(BASE_DIR) + os.sep):
        return jsonify({"error": "Fichier introuvable"}), 404
    if os.path.isfile(full_path):
        resp = send_file(full_path)
        if filepath.endswith('.html'):
            resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        return resp
    return jsonify({"error": "Fichier introuvable"}), 404

# ══════════════════════════════════════════════════════════════════════
# PAGE D'ACCUEIL PUBLIQUE — DONNÉES DYNAMIQUES + LOGIN
# ══════════════════════════════════════════════════════════════════════

DGFD_PLATFORM_DIR = os.path.join(BASE_DIR, 'dgfd_platform_complete')


def _load_accueil_content():
    """Charge le contenu editorial (JSON) de la page d'accueil.
    Le fichier dgfd_platform_complete/data/dgfd_data.json est cree avec
    les valeurs par defaut s'il n'existe pas (meme logique que Streamlit)."""
    if DGFD_PLATFORM_DIR not in sys.path:
        sys.path.insert(0, DGFD_PLATFORM_DIR)
    from data_manager import load_data
    return load_data()


# Les 12 departements officiels du Benin (pour filtrer les zones en texte libre)
DEPARTEMENTS_BENIN = ['ALIBORI', 'ATACORA', 'ATLANTIQUE', 'BORGOU', 'COLLINES',
                      'COUFFO', 'DONGA', 'LITTORAL', 'MONO', 'OUEME', 'PLATEAU', 'ZOU']


def _extract_departements_zone(zone_str):
    """Extrait les departements officiels cites dans une zone (texte libre ou arbre)."""
    if not zone_str or not str(zone_str).strip():
        return set()
    texte = strip_accents(str(zone_str)).upper()
    return {dep for dep in DEPARTEMENTS_BENIN
            if re.search(r'\b' + dep + r'\b', texte)}


def _repartition_projets(rows_actifs, champ, top=None, annee_cloture=None):
    """Repartition (camembert) des projets en cours selon un champ (secteur/partenaire).
    top=None : toutes les valeurs ; sinon top N + "Autres".
    annee_cloture : si fourni, ajoute le nombre de clotures de cette annee par element."""
    from collections import Counter
    comptes = Counter()
    clotures = Counter()
    annee_str = str(annee_cloture) if annee_cloture else None
    for r in rows_actifs:
        val = str(r.get(champ) or '').strip()
        if not val or val.lower() == 'nan':
            val = 'Non précisé'
        comptes[val] += 1
        if annee_str:
            ref = str(r.get('nouvelle_date_cloture') or r.get('date_cloture') or '')[:10]
            if ref[:4] == annee_str:
                clotures[val] += 1
    total = sum(comptes.values())
    if not total:
        return None
    if top and len(comptes) > top:
        items = comptes.most_common(top)
        items.append(('Autres', total - sum(c for _, c in items)))
        top_names = {n for n, _ in items[:-1]}
        autres_clot = sum(c for n, c in clotures.items() if n not in top_names)
    else:
        items = comptes.most_common()
        top_names = None
    palette = ['#0a2540', '#0e7a3a', '#f2c94c', '#2563eb', '#7c3aed', '#dc2626', '#0891b2',
               '#d97706', '#059669', '#db2777', '#4f46e5', '#65a30d', '#0d9488', '#9333ea', '#b45309', '#475569']
    secteurs, pcts, couleurs, infos, effectifs = [], [], [], [], []
    reste = 100
    for i, (nom, nb) in enumerate(items):
        pct = round(nb * 100 / total) if i < len(items) - 1 else reste
        pct = max(0, min(pct, reste))
        reste -= pct
        secteurs.append(nom)
        pcts.append(pct)
        couleurs.append(palette[i % len(palette)])
        effectifs.append(nb)
        if annee_str:
            infos.append(autres_clot if nom == 'Autres' else clotures.get(nom, 0))
    result = {'Secteurs': secteurs, 'Pourcentages': pcts, 'Couleurs': couleurs, 'Total': total,
              'Effectifs': effectifs}
    if annee_str:
        result['Clotures'] = infos
        result['AnneeCloture'] = int(annee_str)
    return result


def _compute_accueil_stats(rows):
    """Calcule les statistiques dynamiques : informations recentes et projets en cours."""
    from datetime import date, timedelta
    aujourd_hui = date.today()
    auj_iso = aujourd_hui.isoformat()
    il_y_a_12_mois = (aujourd_hui - timedelta(days=365)).isoformat()
    dans_12_mois = (aujourd_hui + timedelta(days=365)).isoformat()

    def iso(v):
        return str(v)[:10] if v else ''

    actifs = 0
    clotures_a_venir = 0
    accords_12_mois = 0
    partenaires_actifs = set()
    departements_actifs = set()
    montant_actifs = 0.0
    rows_actifs = []

    for r in rows:
        # Un emprunt obligataire n'est pas un projet : exclu des indicateurs "en cours"
        emprunt = 'EMPRUNT OBLIGATAIRE' in strip_accents(str(r.get('partenaire') or '')).upper()
        # Projet en cours = cloture (apres prorogation sinon initiale) >= aujourd'hui
        ref = iso(r.get('nouvelle_date_cloture')) or iso(r.get('date_cloture'))
        if ref:
            actif = (not emprunt) and ref >= auj_iso
            if actif and ref <= dans_12_mois:
                clotures_a_venir += 1
        else:
            try:
                actif = (not emprunt) and int(r.get('annee_cloture')) >= aujourd_hui.year
            except (TypeError, ValueError):
                actif = not emprunt  # aucune cloture connue = en cours
        if actif:
            actifs += 1
            rows_actifs.append(r)
            part = str(r.get('partenaire') or '').strip()
            if part:
                partenaires_actifs.add(part)
            deps_zone = _extract_departements_zone(r.get('zone'))
            if deps_zone:
                departements_actifs.update(deps_zone)
            else:
                dep = strip_accents(str(r.get('departement') or '').strip()).upper()
                if dep in DEPARTEMENTS_BENIN:
                    departements_actifs.add(dep)
            try:
                montant_actifs += float(r.get('montant_total_fcfa') or 0)
            except (TypeError, ValueError):
                pass
        # Accord signe dans les 12 derniers mois
        d_sig = iso(r.get('date_signature'))
        if d_sig and d_sig >= il_y_a_12_mois:
            accords_12_mois += 1

    milliards = int(round(montant_actifs / 1_000_000_000))
    montant_str = f'{milliards:,}'.replace(',', ' ')

    # Derniers accords signes (4 plus recents)
    tries = sorted(
        [r for r in rows if iso(r.get('date_signature'))],
        key=lambda r: iso(r.get('date_signature')), reverse=True
    )[:4]
    accords_recents = []
    for r in tries:
        d = iso(r.get('date_signature'))
        date_fr = f'{d[8:10]}/{d[5:7]}/{d[:4]}' if len(d) == 10 else d
        objet = str(r.get('objet_accord') or r.get('code_projet') or '').strip()
        if len(objet) > 70:
            objet = objet[:67] + '...'
        accords_recents.append({
            'code': r.get('code_projet', ''),
            'projet': objet,
            'partenaire': str(r.get('partenaire') or '').strip(),
            'date': date_fr,
        })

    stats = {
        'projets_en_cours': {'valeur': actifs, 'label': 'Projets en cours', 'sublabel': "En cours d'exécution", 'icone': '📁'},
        'accords_recents': {'valeur': accords_12_mois, 'label': 'Accords signés', 'sublabel': '12 derniers mois', 'icone': '📄'},
        'financements_en_cours': {'valeur': montant_str, 'label': 'Milliards FCFA', 'sublabel': 'Financements en cours', 'icone': '🪙'},
        'partenaires_actifs': {'valeur': len(partenaires_actifs), 'label': 'Partenaires actifs', 'sublabel': 'Sur les projets en cours', 'icone': '👥'},
        'departements': {'valeur': len(departements_actifs), 'label': 'Départements couverts', 'sublabel': 'Par les projets en cours', 'icone': '🏛️'},
        'clotures_a_venir': {'valeur': clotures_a_venir, 'label': 'Clôtures à venir', 'sublabel': '12 prochains mois', 'icone': '⏳'},
    }
    today_stats = [
        {'icone': '🚀', 'valeur': actifs, 'label': 'Projets en cours'},
        {'icone': '📄', 'valeur': accords_12_mois, 'label': 'Accords signés (12 mois)'},
        {'icone': '🪙', 'valeur': montant_str, 'label': 'Milliards FCFA en cours'},
    ]
    return {'STATS': stats, 'TODAY_STATS': today_stats, 'ACCORDS': accords_recents,
            'REPARTITION': _repartition_projets(rows_actifs, 'secteur_principal', annee_cloture=aujourd_hui.year),
            'REPARTITION_PARTENAIRES': _repartition_projets(rows_actifs, 'partenaire', top=15, annee_cloture=aujourd_hui.year)}


_accueil_cache = {'ts': 0, 'data': None}
_ACCUEIL_CACHE_TTL = 300  # secondes


def _build_accueil_data():
    """Construit le contenu de la page d'accueil (editorial + stats Supabase)."""
    try:
        content = _load_accueil_content()
    except Exception:
        traceback.print_exc()
        content = {}
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('accords_consolides').select(
            'code_projet, partenaire, objet_accord, date_signature, annee_signature, '
            'date_cloture, nouvelle_date_cloture, annee_cloture, '
            'montant_total_fcfa, departement, zone, secteur_principal'
        ).execute()
        dyn = _compute_accueil_stats(resp.data or [])
        content['STATS'] = dyn['STATS']
        content['TODAY_STATS'] = dyn['TODAY_STATS']
        if dyn['ACCORDS']:
            content['ACCORDS'] = dyn['ACCORDS']
        if dyn['REPARTITION']:
            content['REPARTITION'] = dyn['REPARTITION']
        if dyn['REPARTITION_PARTENAIRES']:
            content['REPARTITION_PARTENAIRES'] = dyn['REPARTITION_PARTENAIRES']
        content['source'] = 'supabase'
    except Exception:
        traceback.print_exc()
        content['source'] = 'defaut'  # valeurs par defaut du JSON
    return content


@app.route('/api/accueil/data', methods=['GET'])
def get_accueil_data():
    """Donnees de la page d'accueil: contenu editorial (JSON) + stats reelles (Supabase).
    Resultat mis en cache 5 min pour eviter les lenteurs/cold starts."""
    now = time.time()
    if _accueil_cache['data'] is None or (now - _accueil_cache['ts']) > _ACCUEIL_CACHE_TTL:
        _accueil_cache['data'] = _build_accueil_data()
        _accueil_cache['ts'] = now
    return jsonify(_accueil_cache['data'])


SAFE_OPEN_DATA_FIELDS = [
    'code_projet', 'partenaire', 'objet_accord', 'date_signature', 'annee_signature',
    'date_cloture', 'annee_cloture', 'montant_total_fcfa', 'departement', 'commune', 'zone', 'secteur_principal'
]


def _parse_int(value, default=None):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _clean_text(value):
    return str(value or '').strip()


def _normalize_date(value):
    text = _clean_text(value)
    if len(text) >= 10:
        return text[:10]
    return text


@app.route('/api/open-data', methods=['GET'])
def open_data():
    """Expose un sous-ensemble public de donnees non sensibles, avec filtres et restrictions d'export."""
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('accords_consolides').select(','.join(SAFE_OPEN_DATA_FIELDS)).execute()
        rows = resp.data or []
    except Exception as exc:
        traceback.print_exc()
        return jsonify({
            'source': 'error',
            'generated_at': datetime.utcnow().isoformat() + 'Z',
            'count': 0,
            'records': [],
            'options': {},
            'error': str(exc)
        }), 500

    args = request.args
    filters = {}
    if args.get('partenaire'):
        filters['partenaire'] = str(args.get('partenaire')).strip()
    if args.get('secteur_principal'):
        filters['secteur_principal'] = str(args.get('secteur_principal')).strip()
    if args.get('departement'):
        filters['departement'] = str(args.get('departement')).strip()
    if args.get('commune'):
        filters['commune'] = str(args.get('commune')).strip()
    if args.get('annee_cloture'):
        filters['annee_cloture'] = str(args.get('annee_cloture')).strip()
    if args.get('date_cloture'):
        filters['date_cloture'] = _normalize_date(args.get('date_cloture'))
    search = str(args.get('search') or '').strip()
    limit = min(max(_parse_int(args.get('limit'), 25), 1), 250)
    offset = max(_parse_int(args.get('offset'), 0), 0)
    fmt = str(args.get('format') or 'json').lower()

    def matches(row):
        if search:
            hay = ' '.join([
                str(row.get('code_projet') or ''),
                str(row.get('partenaire') or ''),
                str(row.get('objet_accord') or ''),
                str(row.get('secteur_principal') or ''),
                str(row.get('departement') or ''),
                str(row.get('commune') or ''),
                str(row.get('date_cloture') or ''),
                str(row.get('annee_cloture') or ''),
            ]).lower()
            if search.lower() not in hay:
                return False
        for key, value in filters.items():
            if key == 'date_cloture':
                row_value = _normalize_date(row.get(key))
                if row_value and row_value != str(value).strip():
                    return False
            else:
                row_value = _clean_text(row.get(key) or '').lower()
                if row_value and row_value != str(value).strip().lower():
                    return False
        return True

    filtered_rows = [row for row in rows if matches(row)]

    def _distinct_values(key):
        vals = []
        for row in rows:
            val = _clean_text(row.get(key) or '')
            if val and val not in vals:
                vals.append(val)
        return sorted(vals)

    options = {
        'partenaire': _distinct_values('partenaire'),
        'secteur_principal': _distinct_values('secteur_principal'),
        'departement': _distinct_values('departement'),
        'commune': _distinct_values('commune'),
    }

    if fmt in {'csv', 'xlsx'}:
        if len(filters) < 2 and not search:
            return jsonify({
                'error': 'Les exports CSV/Excel demandent au moins deux filtres combinés ou une recherche ciblée.',
                'message': 'Pour des raisons de confidentialité, les exports sont limités à des requêtes filtrées.'
            }), 400
        if limit > 250:
            return jsonify({'error': 'La limite d’export est fixée à 250 lignes maximum.'}), 400
        export_rows = filtered_rows[offset:offset + limit]
        if fmt == 'csv':
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=SAFE_OPEN_DATA_FIELDS)
            writer.writeheader()
            for row in export_rows:
                writer.writerow({k: row.get(k, '') for k in SAFE_OPEN_DATA_FIELDS})
            output.seek(0)
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = 'attachment; filename="opendata_export.csv"'
            return response
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'open_data'
        ws.append(SAFE_OPEN_DATA_FIELDS)
        for row in export_rows:
            ws.append([row.get(k, '') for k in SAFE_OPEN_DATA_FIELDS])
        wb.save(output)
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename="opendata_export.xlsx"'
        return response

    preview = filtered_rows[offset:offset + limit]
    return jsonify({
        'source': 'supabase',
        'generated_at': datetime.utcnow().isoformat() + 'Z',
        'preview_limit': limit,
        'offset': offset,
        'filters_applied': filters,
        'search': search,
        'count': len(filtered_rows),
        'records': preview,
        'options': options,
        'message': 'Aucune extraction complète n’est fournie. Utiliser plusieurs filtres et des exports limités.'
    })


@app.route('/api/ping')
def api_ping():
    """Point de ping leger pour keep-alive (aucun effet de bord)."""
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════
# VEILLE DES ACCORDS SIGNÉS — détection en ligne gratuite (GDELT/Google News)
# ══════════════════════════════════════════════════════════════════════

@app.route('/api/veille/alertes', methods=['GET'])
@require_auth
def veille_alertes():
    """Liste des alertes de veille (accords signés détectés en ligne)."""
    from db import get_supabase
    sb = get_supabase()
    resp = sb.table('veille_alertes').select('*').order('detecte_le', desc=True).limit(100).execute()
    return jsonify({'data': resp.data or []})


@app.route('/api/veille/scan', methods=['POST'])
@require_auth
def veille_scan():
    """Déclenche un scan manuel de la veille."""
    from services import veille_service
    try:
        n = veille_service.scan_and_notify()
        return jsonify({'message': f'{n} nouvelle(s) alerte(s) enregistrée(s)', 'nouvelles': n})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


def _veille_loop():
    """Scan périodique (toutes les 6 h)."""
    time.sleep(30)  # laisse le serveur démarrer
    while True:
        try:
            from services import veille_service
            n = veille_service.scan_and_notify()
            print(f'[veille] scan terminé : {n} nouvelle(s) alerte(s)')
        except Exception:
            traceback.print_exc()
        time.sleep(6 * 3600)


def _demarrer_veille():
    """Démarre le scan périodique dans un seul worker/processus."""
    if app.debug and os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        return  # processus parent du reloader local
    try:
        fd = os.open(os.path.join(tempfile.gettempdir(), 'sgiad_veille.lock'),
                     os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.close(fd)
    except FileExistsError:
        return  # un autre worker possède déjà la veille
    threading.Thread(target=_veille_loop, daemon=True).start()
    print('[veille] planificateur démarré (scan toutes les 6 h)')


_demarrer_veille()


@app.route('/api/accueil/login', methods=['POST'])
def accueil_login():
    """Connexion depuis la page d'accueil via Supabase Auth
    (l'identifiant est l'email du compte)."""
    try:
        data = request.get_json(force=True) or {}
        identifiant = str(data.get('identifiant', '')).strip()
        mot_de_passe = str(data.get('mot_de_passe', ''))
        if not identifiant or not mot_de_passe:
            return jsonify({'error': 'Identifiant et mot de passe requis'}), 400

        from db import get_supabase
        sb = get_supabase()
        resp = sb.auth.sign_in_with_password({
            'email': identifiant,
            'password': mot_de_passe,
        })
        if not resp or not resp.session:
            return jsonify({'error': 'Identifiant ou mot de passe incorrect'}), 401

        from db import log_session
        log_session(request.remote_addr, resp.user.email or identifiant)

        return jsonify({
            'message': 'Connexion réussie',
            'user': resp.user.email or identifiant,
            'access_token': resp.session.access_token,
            'refresh_token': resp.session.refresh_token,
            'expires_in': resp.session.expires_in,
        })
    except Exception as e:
        traceback.print_exc()
        if 'Invalid login credentials' in str(e):
            return jsonify({'error': 'Identifiant ou mot de passe incorrect'}), 401
        return jsonify({'error': 'Erreur de connexion, veuillez réessayer.'}), 500


# ══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════════════════

def strip_accents(s):
    if not isinstance(s, str):
        return s
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def clean_date(val):
    if pd.isna(val) or str(val).strip() in ('', 'nan'):
        return None
    try:
        return pd.to_datetime(val, dayfirst=True).strftime('%Y-%m-%d')
    except Exception:
        return None

def load_wb_safe(path, retries=3):
    for attempt in range(retries):
        try:
            return load_workbook(path)
        except PermissionError:
            print(f"⚠️  Fichier verrouillé, tentative {attempt+1}/{retries}…")
            time.sleep(1)
    return None

def _safe_error(e):
    """Message d'erreur sur : details techniques seulement en debug local."""
    if os.environ.get('FLASK_DEBUG') == '1':
        return f"{type(e).__name__}: {str(e)}"
    return "Une erreur interne est survenue. Veuillez réessayer."

PLACEHOLDERS = {"-- sélectionner --", "-- selectionner --", "tous", "tout", ""}

def normalize_filter_values(values):
    flat = []
    def _add(v):
        if v is None:
            return
        if isinstance(v, (list, tuple, set)):
            for sub in v: _add(sub)
        elif isinstance(v, dict):
            _add(v.get('value') or v.get('label') or str(v))
        else:
            s = str(v).strip()
            if s and s.lower() not in PLACEHOLDERS:
                flat.append(s)
    _add(values)
    seen, result = set(), []
    for x in flat:
        if x not in seen:
            seen.add(x); result.append(x)
    return result

# ══════════════════════════════════════════════════════════════════════════════
# SESSION / EN-TÊTE ACCUEIL (Supabase)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/session/info', methods=['GET'])
@require_auth
def get_session_info():
    try:
        from db import get_session_stats
        user = getattr(request, 'current_user', {}) or {}
        result = get_session_stats(request.remote_addr)
        if user.get('email'):
            result['user'] = user['email']
        return jsonify(result)
    except Exception as e:
        # Fallback: repondre avec des valeurs par defaut si Supabase indisponible
        traceback.print_exc()
        return jsonify({
            "ip": request.remote_addr,
            "derniere_connexion": None,
            "total_connexions": 0
        })

# ══════════════════════════════════════════════════════════════════════════════
# LOCALISATION HIÉRARCHIQUE (Supabase)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/localisation/hierarchie', methods=['GET'])
@require_auth
def get_localisation_hierarchie():
    try:
        from db import get_localisation_tree
        return jsonify(get_localisation_tree())
    except Exception as e:
        # Fallback Excel si Supabase indisponible
        try:
            tree = _get_localisation_from_excel()
            return jsonify(tree)
        except Exception as e2:
            traceback.print_exc()
            return jsonify({"error": _safe_error(e2)}), 500


@app.route('/api/localisation/reload', methods=['POST'])
@require_auth
def reload_localisation_tree():
    try:
        from db import get_localisation_tree
        get_localisation_tree(force_reload=True)
        return jsonify({"message": "Arbre de localisation rechargé"}), 200
    except Exception as e:
        return jsonify({"error": _safe_error(e)}), 500


def _get_localisation_from_excel():
    """Fallback: lit la localisation depuis Excel si Supabase echoue."""
    import re as _re
    def _clean_village(v):
        return _re.sub(r'^\d+\.\s*', '', str(v).strip())

    df = pd.read_excel(BANQUE_FILE, sheet_name='Localisation', dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    tree = {}
    for _, row in df.iterrows():
        dep = str(row['Département']).strip()
        com = str(row['Commune']).strip()
        arr = str(row['Arrondissement']).strip()
        vil = _clean_village(row['Village ou quartier de Ville'])
        tree.setdefault(dep, {}).setdefault(com, {}).setdefault(arr, [])
        tree[dep][com][arr].append(vil)
    return tree


# ══════════════════════════════════════════════════════════════════════════════
# MODULE BANQUE DE PROJETS (Projets - reste sur Excel pour l'instant)
# ══════════════════════════════════════════════════════════════════════════════

def decode_localisation(s):
    def split_top(text, sep):
        parts, depth, cur = [], 0, ""
        for ch in text:
            if ch == '(':
                depth += 1; cur += ch
            elif ch == ')':
                depth -= 1; cur += ch
            elif ch == sep and depth == 0:
                parts.append(cur.strip()); cur = ""
            else:
                cur += ch
        if cur.strip():
            parts.append(cur.strip())
        return parts

    def parse_node(item):
        item = item.strip()
        if item.endswith(')') and '(' in item:
            idx = item.index('(')
            name = item[:idx].strip()
            inner = item[idx + 1:-1]
            children = {}
            for child_item in split_top(inner, ','):
                cname, cval = parse_node(child_item)
                children[cname] = cval
            return name, children
        return item, {}

    if not s or not str(s).strip():
        return {}
    s = str(s).strip().rstrip('.').strip()
    result = {}
    for dep_item in split_top(s, ';'):
        name, val = parse_node(dep_item)
        result[name] = val
    return result

def flatten_localisation_names(tree):
    names = set()
    def walk(node):
        for name, children in node.items():
            names.add(name)
            if children:
                walk(children)
    walk(tree)
    return names

def zone_intervention_matches(zone_str, filter_set_upper):
    if not zone_str or str(zone_str).strip() == '':
        return False
    names = flatten_localisation_names(decode_localisation(zone_str))
    names_upper = {n.strip().upper() for n in names}
    return bool(names_upper & filter_set_upper)

def derive_localisation_columns(zone_str):
    tree = decode_localisation(zone_str)
    deps, coms, arrs, vils = [], [], [], []
    for dep, dep_children in tree.items():
        if dep and dep not in deps:
            deps.append(dep)
        for com, com_children in (dep_children or {}).items():
            if com and com not in coms:
                coms.append(com)
            for arr, arr_children in (com_children or {}).items():
                if arr and arr not in arrs:
                    arrs.append(arr)
                for vil in (arr_children or {}).keys():
                    if vil and vil not in vils:
                        vils.append(vil)
    return {
        'Departement': ', '.join(deps),
        'Commune': ', '.join(coms),
        'Arrondissement': ', '.join(arrs),
        'Village_Quartier': ', '.join(vils),
    }


@app.route('/api/banque/projets/list', methods=['GET'])
@require_auth
def get_projets_banque():
    try:
        if not os.path.exists(BANQUE_FILE):
            return jsonify({"error": "Fichier Banque_Projets.xlsx introuvable"}), 404
        df = pd.read_excel(BANQUE_FILE, sheet_name='Projets', dtype=str, header=1)
        df = df.dropna(subset=['Code_Projet'])
        df = df[df['Code_Projet'].astype(str).str.strip() != '']

        zone_params = request.args.getlist('zone')
        if zone_params and 'Zone_intervention' in df.columns:
            filter_set_upper = {z.strip().upper() for z in zone_params if z.strip()}
            if filter_set_upper:
                mask = df['Zone_intervention'].apply(
                    lambda s: zone_intervention_matches(s, filter_set_upper))
                df = df.loc[mask].copy()

        date_cols = ['Date_Approbation','Date_Signature','Date_Vigueur','Date_Demarrage',
                     'Date_Cloture','Date_Prorogation','Nouvelle_Date_Cloture',
                     'Date_Soumission','Date_Validation']
        for col in date_cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_date)
        data = json.loads(df.to_json(orient='records', force_ascii=False))
        cleaned = [{k: v for k, v in r.items() if v is not None and str(v).strip() != ''} for r in data]

        for r in cleaned:
            zone_str = r.get('Zone_intervention')
            if zone_str and str(zone_str).strip() and not r.get('Departement'):
                for field, value in derive_localisation_columns(zone_str).items():
                    if value:
                        r[field] = value

        return jsonify(cleaned)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/projets/add', methods=['POST'])
@require_auth
def add_projet_banque():
    try:
        data = request.get_json(force=True) or {}
        zone_str = data.get('Zone_intervention') or data.get('Zone', '')
        if zone_str and str(zone_str).strip():
            for field, value in derive_localisation_columns(zone_str).items():
                if not data.get(field) and value:
                    data[field] = value
        wb = load_wb_safe(BANQUE_FILE)
        if wb is None:
            return jsonify({"error": "Fichier verrouillé. Fermez-le et réessayez."}), 503
        ws = wb['Projets']
        row_data = [
            data.get('Code_Projet'), data.get('Intitule_Projet'), data.get('Annee_Collecte'),
            data.get('Categorie'), data.get('Envergure'), data.get('Secteur_Activite_1'),
            data.get('Secteur_Activite_2'), data.get('Secteur_Activite_3'), data.get('Sous_Secteur'),
            data.get('Tutelle'), data.get('Agence_Exécution'), data.get('Contexte_Justification'),
            data.get('Problematique'), data.get('Objectif_Global'), data.get('Objectifs_Specifiques'),
            data.get('Principaux_Resultats'), data.get('Composantes_Projet'), data.get('Activites_Principales'),
            data.get('Beneficiaires_Directs'), data.get('Beneficiaires_Indirects'), data.get('Departement'),
            data.get('Commune'), data.get('Arrondissement'), data.get('Village_Quartier'),
            data.get('Coordonnees_GPS'), data.get('Date_Approbation'), data.get('Date_Signature'),
            data.get('Date_Vigueur'), data.get('Date_Demarrage'), data.get('Date_Cloture'),
            data.get('Duree_Initiale_Mois'), data.get('Est_Proroge'), data.get('Date_Prorogation'),
            data.get('Motif_Prorogation'), data.get('Nouvelle_Date_Cloture'), data.get('Nb_Prorogations'),
            data.get('Cout_Total_USD'), data.get('Devise'), data.get('Taux_Change'),
            data.get('Cout_Total_FCFA'), data.get('Milliards_FCFA'),
            data.get('Contrepartie_BN_USD'), data.get('Contrepartie_BN_FCFA'),
            data.get('Don_Attendu_FCFA'), data.get('Pret_Attendu_FCFA'), data.get('PPP_Attendu_FCFA'),
            data.get('Contribution_Benef_FCFA'), data.get('Gap_Financement'), data.get('Pourcentage_Acquis'),
            data.get('Source_Financement'), data.get('APD_Oui_Non'), data.get('Type_Financement'),
            data.get('Contributeur'), data.get('Sigle_Partenaire'), data.get('Type_Contributeur'),
            data.get('Co_Partenaire_1'), data.get('Co_Partenaire_2'), data.get('Co_Partenaire_3'),
            data.get('Co_Partenaire_4'), data.get('Co_Partenaire_5'), data.get('Co_Partenaire_6'),
            data.get('Co_Partenaire_7'), data.get('Statut_Projet'), data.get('Secteur_PAG'),
            data.get('Pilier_PAG'), data.get('Axe_PAG'), data.get('ODD'), data.get('Cible_ODD'),
            data.get('Pilier_National'), data.get('Resp_Banque_Nom'), data.get('Resp_Banque_Tel1'),
            data.get('Resp_Banque_Tel2'), data.get('Resp_Banque_Email'), data.get('Coord_Projet_Nom'),
            data.get('Coord_Projet_Tel1'), data.get('Coord_Projet_Tel2'), data.get('Coord_Projet_Email'),
            data.get('RAF_Nom'), data.get('RAF_Tel1'), data.get('RAF_Tel2'), data.get('RAF_Email'),
            data.get('Resp_SE_Nom'), data.get('Resp_SE_Tel1'), data.get('Resp_SE_Tel2'), data.get('Resp_SE_Email'),
            data.get('Niveau_Maturite'), data.get('Etude_Faisabilite'), data.get('Etude_Environnementale'),
            data.get('Plan_Affaires'), data.get('Dossier_Technique'), data.get('Categorie_Env'),
            data.get('Theme_Genre'), data.get('Theme_Inclusion'), data.get('Theme_Climat'),
            data.get('Theme_Emploi_Jeunes'), data.get('Risques_Identifies'), data.get('Mesures_Attenuation'),
            data.get('Niveau_Risque'), data.get('Date_Soumission'), data.get('Avis_Technique'),
            data.get('Avis_DGFD'), data.get('Avis_Ministere'), data.get('Decision_Comite'),
            data.get('Date_Validation'), data.get('Note_Conceptuelle'), data.get('Cadre_Logique'),
            data.get('Budget_Detaille'), data.get('Cartographie'), data.get('Lettre_Demande'),
            data.get('Zone_intervention'),
            data.get('Modalite_Intervention'), data.get('Nature_Financement'),
            data.get('Instrument_Financement'), data.get('Source_Verifiable'),
            data.get('Montant_Devise_Origine'),
        ]
        ws.append(row_data)
        wb.save(BANQUE_FILE); wb.close()
        return jsonify({"message": "Projet ajouté avec succès", "code": data.get('Code_Projet')}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/projets/update/<code>', methods=['PUT'])
@require_auth
def update_projet_banque(code):
    try:
        code = unquote(code).strip()
        updated_data = request.get_json(force=True) or {}
        zone_str = updated_data.get('Zone_intervention') or updated_data.get('Zone', '')
        if zone_str and str(zone_str).strip():
            for field, value in derive_localisation_columns(zone_str).items():
                if not updated_data.get(field) and value:
                    updated_data[field] = value
        wb = load_wb_safe(BANQUE_FILE)
        if wb is None:
            return jsonify({"error": "Fichier verrouillé. Fermez-le et réessayez."}), 503
        ws = wb['Projets']
        row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if str(row[0]).strip() == code:
                row_idx = i; break
        if not row_idx:
            return jsonify({"error": f"Projet {code} non trouvé"}), 404
        col_mapping = {
            'Code_Projet':1,'Intitule_Projet':2,'Annee_Collecte':3,'Categorie':4,'Envergure':5,
            'Secteur_Activite_1':6,'Secteur_Activite_2':7,'Secteur_Activite_3':8,'Sous_Secteur':9,
            'Tutelle':10,'Agence_Exécution':11,'Contexte_Justification':12,'Problematique':13,
            'Objectif_Global':14,'Objectifs_Specifiques':15,'Principaux_Resultats':16,
            'Composantes_Projet':17,'Activites_Principales':18,'Beneficiaires_Directs':19,
            'Beneficiaires_Indirects':20,'Departement':21,'Commune':22,'Arrondissement':23,
            'Village_Quartier':24,'Coordonnees_GPS':25,'Date_Approbation':26,'Date_Signature':27,
            'Date_Vigueur':28,'Date_Demarrage':29,'Date_Cloture':30,'Duree_Initiale_Mois':31,
            'Est_Proroge':32,'Date_Prorogation':33,'Motif_Prorogation':34,'Nouvelle_Date_Cloture':35,
            'Nb_Prorogations':36,'Cout_Total_USD':37,'Devise':38,'Taux_Change':39,
            'Cout_Total_FCFA':40,'Milliards_FCFA':41,
            'Contrepartie_BN_USD':42,'Contrepartie_BN_FCFA':43,'Don_Attendu_FCFA':44,'Pret_Attendu_FCFA':45,
            'PPP_Attendu_FCFA':46,'Contribution_Benef_FCFA':47,'Gap_Financement':48,'Pourcentage_Acquis':49,
            'Source_Financement':50,
            'APD_Oui_Non':51,'Type_Financement':52,'Contributeur':53,'Sigle_Partenaire':54,
            'Type_Contributeur':55,'Co_Partenaire_1':56,'Co_Partenaire_2':57,'Co_Partenaire_3':58,
            'Co_Partenaire_4':59,'Co_Partenaire_5':60,'Co_Partenaire_6':61,'Co_Partenaire_7':62,
            'Statut_Projet':63,'Secteur_PAG':64,'Pilier_PAG':65,'Axe_PAG':66,'ODD':67,
            'Cible_ODD':68,'Pilier_National':69,'Resp_Banque_Nom':70,'Resp_Banque_Tel1':71,
            'Resp_Banque_Tel2':72,'Resp_Banque_Email':73,'Coord_Projet_Nom':74,'Coord_Projet_Tel1':75,
            'Coord_Projet_Tel2':76,'Coord_Projet_Email':77,'RAF_Nom':78,'RAF_Tel1':79,
            'RAF_Tel2':80,'RAF_Email':81,'Resp_SE_Nom':82,'Resp_SE_Tel1':83,'Resp_SE_Tel2':84,
            'Resp_SE_Email':85,'Niveau_Maturite':86,'Etude_Faisabilite':87,
            'Etude_Environnementale':88,'Plan_Affaires':89,'Dossier_Technique':90,
            'Categorie_Env':91,'Theme_Genre':92,'Theme_Inclusion':93,'Theme_Climat':94,
            'Theme_Emploi_Jeunes':95,'Risques_Identifies':96,'Mesures_Attenuation':97,
            'Niveau_Risque':98,'Date_Soumission':99,'Avis_Technique':100,'Avis_DGFD':101,
            'Avis_Ministere':102,'Decision_Comite':103,'Date_Validation':104,
            'Note_Conceptuelle':105,'Cadre_Logique':106,'Budget_Detaille':107,
            'Cartographie':108,'Lettre_Demande':109,
            'Zone_intervention':110,
            'Modalite_Intervention':111,'Nature_Financement':112,'Instrument_Financement':113,
            'Source_Verifiable':114,'Montant_Devise_Origine':115,
        }
        numeric_fields = {'Cout_Total_USD','Taux_Change','Contrepartie_BN_USD','Don_Attendu_FCFA',
                          'Pret_Attendu_FCFA','PPP_Attendu_FCFA','Contribution_Benef_FCFA',
                          'Niveau_Maturite','Nb_Prorogations','Montant_Devise_Origine',
                          'Cout_Total_FCFA','Milliards_FCFA','Contrepartie_BN_FCFA','Gap_Financement'}
        for field, col_idx in col_mapping.items():
            if field in updated_data and updated_data[field] is not None:
                val = updated_data[field]
                if field in numeric_fields:
                    try:
                        val = float(val) if val else None
                    except (ValueError, TypeError):
                        pass
                ws.cell(row=row_idx, column=col_idx, value=val)
        wb.save(BANQUE_FILE); wb.close()
        return jsonify({"message": "Projet mis à jour"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/projets/delete/<code>', methods=['DELETE'])
@require_auth
def delete_projet_banque(code):
    try:
        code = unquote(code).strip()
        wb = load_wb_safe(BANQUE_FILE)
        if wb is None:
            return jsonify({"error": "Fichier verrouillé."}), 503
        ws = wb['Projets']
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if str(row[0]).strip() == code:
                ws.delete_rows(i, 1); wb.save(BANQUE_FILE); wb.close()
                return jsonify({"message": "Projet supprimé"}), 200
        wb.close()
        return jsonify({"error": "Projet non trouvé"}), 404
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/parametres', methods=['GET'])
@require_auth
def get_parametres_banque():
    """Retourne les parametres (Supabase avec fallback Excel).
    Supporte ?categorie=XXX pour filtrer par categorie."""
    categorie = request.args.get('categorie', '').strip() or None
    try:
        from db import get_parametres
        params = get_parametres(categorie=categorie)
        return jsonify(params)
    except Exception:
        # Fallback Excel
        try:
            df = pd.read_excel(BANQUE_FILE, sheet_name='Paramètres', dtype=str)
            params = {}
            for col in df.columns:
                params[col] = [x for x in df[col].dropna().unique().tolist() if str(x).strip() != '']
            if categorie and categorie in params:
                return jsonify(params[categorie])
            return jsonify(params)
        except Exception as e2:
            return jsonify({"error": _safe_error(e2)}), 500


@app.route('/api/banque/parametres/add', methods=['POST'])
@require_auth
def add_parametre_banque():
    try:
        data = request.json
        colonne = str(data.get('colonne', '')).strip()
        nouvelle_valeur = str(data.get('valeur', '')).strip()
        if not colonne or not nouvelle_valeur:
            return jsonify({"error": "Colonne et valeur requises"}), 400

        # Supabase
        try:
            from db import get_supabase
            sb = get_supabase()
            # Verifier doublon (case-insensitive, requête ciblée)
            escaped = nouvelle_valeur.replace('%', '\\%').replace('_', '\\_')
            existing = sb.table('parametres').select('id').eq(
                'categorie', colonne).ilike('valeur', escaped).limit(1).execute()
            if existing.data:
                return jsonify({"message": f"Cette valeur existe déjà : '{nouvelle_valeur}'", "doublon": True}), 409
            # Inserer
            sb.table('parametres').insert({
                'categorie': colonne,
                'valeur': nouvelle_valeur,
                'ordre': 999,
            }).execute()
            return jsonify({"message": "Valeur ajoutée", "colonne": colonne, "valeur": nouvelle_valeur}), 201
        except Exception as sb_err:
            traceback.print_exc()
            return jsonify({"error": _safe_error(sb_err)}), 500

        # Fallback Excel
        wb = load_wb_safe(BANQUE_FILE)
        if wb is None:
            return jsonify({"error": "Fichier verrouillé."}), 503
        ws = wb['Paramètres']
        col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value and str(c.value).strip().upper() == colonne.upper()), None)
        if not col_idx:
            wb.close(); return jsonify({"error": f"Colonne '{colonne}' introuvable"}), 404
        existing = {str(r[0]).strip().upper() for r in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True) if r[0]}
        if nouvelle_valeur.upper() in existing:
            wb.close(); return jsonify({"message": "Doublon", "doublon": True}), 409
        next_row = 2
        while ws.cell(row=next_row, column=col_idx).value not in (None, ''):
            next_row += 1
        ws.cell(row=next_row, column=col_idx, value=nouvelle_valeur)
        wb.save(BANQUE_FILE); wb.close()
        return jsonify({"message": "Valeur ajoutée", "colonne": colonne, "valeur": nouvelle_valeur}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/parametres/delete', methods=['POST'])
@require_auth
def delete_parametre_banque():
    try:
        data = request.get_json(force=True) or {}
        colonne = str(data.get('colonne', '')).strip()
        valeur  = str(data.get('valeur', '')).strip()
        if not colonne or not valeur:
            return jsonify({"error": "Colonne et valeur requises"}), 400

        from db import get_supabase
        sb = get_supabase()

        # Mapping nom affiche -> colonne DB
        DISPLAY_TO_DB = {
            'Secteur principal': 'secteur_principal',
            'SOUS SECTEUR': 'sous_secteur',
            "Modalit\u00e9 d'intervention": 'modalite_intervention',
            'Nature (Pr\u00eat/Don/Mixte)': 'nature_pret_don_mixte',
            'Devise': 'devise',
            'APD Oui/non': 'apd_oui_non',
            'TYPE DE FINANCEMENT': 'type_financement',
            'TYPE DE CONTRIBUTEUR': 'type_contributeur',
            'Instrument de financement': 'instrument_financement',
            'Cible_ODD': 'cible_odd',
            'AXE PAG': 'axe_pag',
            'PILIER PAG': 'pilier_pag',
            'ODD': 'odd',
            'Pilier': 'pilier',
            'STATUT (en cours, achev\u00e9, en approbation)': 'statut',
            'Tutelle': 'tutelle',
            'Agence Ex\u00e9cution': 'agence_execution',
            'Partenaire': 'partenaire',
        }

        # Verifier si la valeur est utilisee dans un projet
        db_col = DISPLAY_TO_DB.get(colonne)
        if db_col:
            try:
                check = sb.table('accords_consolides').select('id').eq(db_col, valeur).limit(1).execute()
                if check.data and len(check.data) > 0:
                    return jsonify({
                        "error": f"Impossible de supprimer '{valeur}' : cette valeur est utilis\u00e9e dans un projet existant. Veuillez d'abord supprimer ou modifier le projet concern\u00e9.",
                        "used_in_project": True
                    }), 409
            except Exception:
                pass  # Si la verification echoue, on proceed quand meme

        # Supprimer de parametres
        sb.table('parametres').delete().eq(
            'categorie', colonne).eq('valeur', valeur).execute()
        return jsonify({"message": "Valeur supprim\u00e9e"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/parametres/update', methods=['POST'])
@require_auth
def update_parametre_banque():
    try:
        data = request.json
        colonne = str(data.get('colonne', '')).strip()
        ancienne_valeur = str(data.get('ancienne_valeur', '')).strip()
        nouvelle_valeur = str(data.get('nouvelle_valeur', '')).strip()
        if not colonne or not ancienne_valeur or not nouvelle_valeur:
            return jsonify({"error": "Colonne, ancienne et nouvelle valeurs requises"}), 400

        # Supabase
        try:
            from db import get_supabase
            sb = get_supabase()
            # Verifier doublon
            existing = sb.table('parametres').select('id').eq(
                'categorie', colonne).eq('valeur', nouvelle_valeur).limit(1).execute()
            if existing.data:
                return jsonify({"message": "Doublon", "doublon": True}), 409
            # Update
            sb.table('parametres').update({
                'valeur': nouvelle_valeur,
            }).eq('categorie', colonne).eq('valeur', ancienne_valeur).execute()
            return jsonify({"message": "Valeur modifiée", "colonne": colonne}), 200
        except Exception as sb_err:
            traceback.print_exc()
            return jsonify({"error": _safe_error(sb_err)}), 500

        # Fallback Excel
        wb = load_wb_safe(BANQUE_FILE)
        if wb is None:
            return jsonify({"error": "Fichier verrouillé."}), 503
        ws = wb['Paramètres']
        col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value and str(c.value).strip().upper() == colonne.upper()), None)
        if not col_idx:
            wb.close(); return jsonify({"error": f"Colonne '{colonne}' introuvable"}), 404
        cell_to_update = next(
            (ws.cell(row=i, column=col_idx) for i in range(2, ws.max_row + 1)
             if ws.cell(row=i, column=col_idx).value and str(ws.cell(row=i, column=col_idx).value).strip().upper() == ancienne_valeur.upper()), None)
        if not cell_to_update:
            wb.close(); return jsonify({"message": "Valeur non trouvée"}), 404
        cell_to_update.value = nouvelle_valeur
        wb.save(BANQUE_FILE); wb.close()
        return jsonify({"message": "Valeur modifiée", "colonne": colonne}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


# ============================================
# DOCUMENTS ASSOCIES AUX PROJETS (stockés dans Supabase)
# ============================================
@app.route('/api/banque/documents', methods=['GET'])
@require_auth
def list_documents_projet():
    """Liste les métadonnées des documents d'un projet (sans le contenu)."""
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({"error": "Paramètre code requis"}), 400
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('documents_projets').select(
            'id, code_projet, categorie, nom_fichier, type_mime, taille_octets, ajoute_le'
        ).eq('code_projet', code).order('categorie').execute()
        return jsonify(resp.data or [])
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/documents/upload', methods=['POST'])
@require_auth
def upload_document_projet():
    """Reçoit un fichier (multipart) et le stocke en base64 dans Supabase."""
    try:
        file = request.files.get('file')
        code_projet = str(request.form.get('code_projet', '')).strip()
        categorie = str(request.form.get('categorie', 'Autre')).strip() or 'Autre'
        if not file or not code_projet:
            return jsonify({"error": "Fichier et code_projet requis"}), 400

        # Quick win 2 : liste blanche extensions + types MIME autorises
        ALLOWED_UPLOADS = {
            '.pdf': 'application/pdf',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.txt': 'text/plain',
        }
        ext = os.path.splitext(secure_filename(file.filename or ''))[1].lower()
        mime = (file.mimetype or '').split(';')[0].strip().lower()
        if ext not in ALLOWED_UPLOADS:
            return jsonify({"error": "Type de fichier non autorisé"}), 400
        if mime not in list(ALLOWED_UPLOADS.values()) + ['application/octet-stream']:
            return jsonify({"error": "Type de fichier non autorisé"}), 400

        content = file.read()
        if len(content) > 10 * 1024 * 1024:
            return jsonify({"error": "Fichier trop volumineux (max 10 Mo)"}), 413
        import base64
        from db import get_supabase
        sb = get_supabase()
        sb.table('documents_projets').insert({
            'code_projet': code_projet,
            'categorie': categorie,
            'nom_fichier': secure_filename(file.filename or 'document.pdf'),
            'type_mime': file.mimetype or 'application/pdf',
            'taille_octets': len(content),
            'contenu_base64': base64.b64encode(content).decode('ascii'),
        }).execute()
        return jsonify({"message": "Document ajouté", "code_projet": code_projet}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/documents/<int:doc_id>/download', methods=['GET'])
@require_auth
def download_document_projet(doc_id):
    """Retourne le contenu du document pour téléchargissement/visualisation."""
    try:
        import base64
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('documents_projets').select(
            'nom_fichier, type_mime, contenu_base64'
        ).eq('id', doc_id).limit(1).execute()
        if not resp.data:
            return jsonify({"error": "Document introuvable"}), 404
        doc = resp.data[0]
        data = base64.b64decode(doc.get('contenu_base64') or '')
        return send_file(
            BytesIO(data),
            mimetype=doc.get('type_mime') or 'application/pdf',
            as_attachment=False,
            download_name=doc.get('nom_fichier') or 'document.pdf'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/documents/<int:doc_id>', methods=['DELETE'])
@require_auth
def delete_document_projet(doc_id):
    try:
        from db import get_supabase
        sb = get_supabase()
        sb.table('documents_projets').delete().eq('id', doc_id).execute()
        return jsonify({"message": "Document supprimé"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/projets/export', methods=['POST'])
@require_auth
def export_projets():
    try:
        req_data = request.get_json(force=True) or {}
        raw_filters = req_data.get('filters', {})
        selected_columns = req_data.get('selected_columns', [])
        if not os.path.exists(BANQUE_FILE):
            return jsonify({"error": "Fichier introuvable"}), 404
        df = pd.read_excel(BANQUE_FILE, sheet_name='Projets', dtype=str, header=1).reset_index(drop=True)
        if df.columns.duplicated().any():
            df = df.loc[:, ~df.columns.duplicated()].copy()
        df.columns = [str(c).strip() for c in df.columns]
        for col_name, raw_values in raw_filters.items():
            col_name = str(col_name).strip()
            if col_name not in df.columns: continue
            filter_list = normalize_filter_values(raw_values)
            if not filter_list: continue

            if col_name == 'Zone_intervention':
                filter_set_upper = {f.strip().upper() for f in filter_list}
                mask = df[col_name].apply(lambda s: zone_intervention_matches(s, filter_set_upper))
                df = df.loc[mask].copy()
                continue

            serie = df[col_name]
            if isinstance(serie, pd.DataFrame): serie = serie.iloc[:, 0]
            df = df.loc[serie.astype(str).str.strip().isin(filter_list)].copy()
        if len(df) == 0:
            return jsonify({"error": "Aucun résultat"}), 404

        if selected_columns:
            valid_cols = []
            for c in selected_columns:
                name = str(c.get('value') if isinstance(c, dict) else c).strip()
                if name in df.columns and name not in valid_cols:
                    valid_cols.append(name)
            if valid_cols: df = df[valid_cols].copy()
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Export')
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'Rapport_Projets_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# SUIVI TRIMESTRIEL (Supabase avec fallback Excel)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/banque/suivi/list', methods=['GET'])
@require_auth
def get_suivi_trimestriel():
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('suivi_trimestriel').select('*').order('code_projet').execute()
        return jsonify(resp.data)
    except Exception:
        # Fallback Excel
        try:
            df = pd.read_excel(BANQUE_FILE, sheet_name='Suivi_Trimestriel', dtype=str)
            df = df.dropna(how='all')
            return jsonify(json.loads(df.to_json(orient='records', force_ascii=False)))
        except Exception as e2:
            traceback.print_exc()
            return jsonify({"error": _safe_error(e2)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# SECTEURS <-> SOUS-SECTEURS
# ══════════════════════════════════════════════════════════════════════════════

_secteurs_tree_cache = None

def _norm_secteur(s):
    s = str(s or '').strip().upper().replace('&', 'ET')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s).strip()

def get_secteurs_sous_secteurs(force_reload=False):
    global _secteurs_tree_cache
    if _secteurs_tree_cache is not None and not force_reload:
        return _secteurs_tree_cache

    display, tree, sub_display = {}, {}, {}

    # Essayer Supabase d'abord
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('accords_consolides').select('secteur_principal, sous_secteur').execute()
        for row in resp.data:
            sec = str(row.get('secteur_principal') or '').strip()
            if not sec or sec.lower() == 'nan':
                continue
            ksec = _norm_secteur(sec)
            display.setdefault(ksec, sec)
            tree.setdefault(ksec, set())
            sous = str(row.get('sous_secteur') or '').strip()
            if sous and sous.lower() != 'nan':
                for part in re.split(r'\s*/\s*', sous):
                    part = part.strip()
                    if part:
                        ksub = _norm_secteur(part)
                        sub_display.setdefault(ksub, part)
                        tree[ksec].add(ksub)
    except Exception:
        # Fallback Excel
        try:
            df = pd.read_excel(BANQUE_FILE, sheet_name='Projets', dtype=str, header=1)
            for _, row in df.iterrows():
                sec = str(row.get('Secteur_Activite_1') or '').strip()
                if not sec or sec.lower() == 'nan':
                    continue
                ksec = _norm_secteur(sec)
                display.setdefault(ksec, sec)
                tree.setdefault(ksec, set())
                sous = str(row.get('Sous_Secteur') or '').strip()
                if sous and sous.lower() != 'nan':
                    for part in re.split(r'\s*/\s*', sous):
                        part = part.strip()
                        if part:
                            ksub = _norm_secteur(part)
                            sub_display.setdefault(ksub, part)
                            tree[ksec].add(ksub)
        except Exception:
            traceback.print_exc()

    result = {}
    for ksec in sorted(display.keys()):
        result[display[ksec]] = sorted(
            (sub_display[k] for k in tree.get(ksec, set())),
            key=lambda x: str(x).upper()
        )
    _secteurs_tree_cache = result
    return result


@app.route('/api/banque/secteurs_sous_secteurs', methods=['GET'])
@require_auth
def get_secteurs_sous_secteurs_route():
    try:
        return jsonify(get_secteurs_sous_secteurs())
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/secteurs_sous_secteurs/reload', methods=['POST'])
@require_auth
def reload_secteurs_sous_secteurs():
    try:
        get_secteurs_sous_secteurs(force_reload=True)
        return jsonify({"message": "Mapping secteurs/sous-secteurs rechargé"}), 200
    except Exception as e:
        return jsonify({"error": _safe_error(e)}), 500


# ── CRUD: secteur_sous_secteur (table de relation explicite) ──

@app.route('/api/banque/secteur_sous_secteur/list', methods=['GET'])
@require_auth
def list_secteur_sous_secteur():
    """Retourne toutes les relations secteur -> sous-secteur."""
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('secteur_sous_secteur').select('*').order('secteur').order('sous_secteur').execute()
        return jsonify(resp.data or [])
    except Exception as e:
        # Table may not exist yet - return empty
        if 'relation' in str(e) and 'does not exist' in str(e):
            return jsonify([])
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/secteur_sous_secteur/by_secteur', methods=['GET'])
@require_auth
def get_sous_secteurs_by_secteur():
    """Retourne les sous-secteurs liés à un secteur donné."""
    secteur = request.args.get('secteur', '').strip()
    if not secteur:
        return jsonify({"error": "Parametre 'secteur' requis"}), 400
    try:
        from db import get_supabase
        sb = get_supabase()
        resp = sb.table('secteur_sous_secteur').select('sous_secteur').eq('secteur', secteur).order('sous_secteur').execute()
        sous_secteurs = [r['sous_secteur'] for r in (resp.data or [])]
        return jsonify(sous_secteurs)
    except Exception as e:
        if 'relation' in str(e) and 'does not exist' in str(e):
            return jsonify([])
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/secteur_sous_secteur/add', methods=['POST'])
@require_auth
def add_secteur_sous_secteur():
    """Ajoute une relation secteur -> sous-secteur."""
    try:
        data = request.json
        secteur = str(data.get('secteur', '')).strip()
        sous_secteur = str(data.get('sous_secteur', '')).strip()
        if not secteur or not sous_secteur:
            return jsonify({"error": "Secteur et sous-secteur requis"}), 400

        from db import get_supabase
        sb = get_supabase()
        # Verifier doublon
        existing = sb.table('secteur_sous_secteur').select('id').eq(
            'secteur', secteur).eq('sous_secteur', sous_secteur).execute()
        if existing.data:
            return jsonify({"message": f"Ce sous-secteur est déjà lié à '{secteur}'", "doublon": True}), 409
        # Inserer
        sb.table('secteur_sous_secteur').insert({
            'secteur': secteur,
            'sous_secteur': sous_secteur,
        }).execute()
        # Recharger le cache
        get_secteurs_sous_secteurs(force_reload=True)
        return jsonify({"message": "Sous-secteur lié", "secteur": secteur, "sous_secteur": sous_secteur}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/secteur_sous_secteur/delete', methods=['POST'])
@require_auth
def delete_secteur_sous_secteur():
    """Supprime une relation secteur -> sous-secteur."""
    try:
        data = request.json
        secteur = str(data.get('secteur', '')).strip()
        sous_secteur = str(data.get('sous_secteur', '')).strip()
        if not secteur or not sous_secteur:
            return jsonify({"error": "Secteur et sous-secteur requis"}), 400

        from db import get_supabase
        sb = get_supabase()
        sb.table('secteur_sous_secteur').delete().eq(
            'secteur', secteur).eq('sous_secteur', sous_secteur).execute()
        # Recharger le cache
        get_secteurs_sous_secteurs(force_reload=True)
        return jsonify({"message": "Liaison supprimée"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# MODULE ACCORDS (Banque_Accords_V2.xlsx) - NON MIGRE
# ══════════════════════════════════════════════════════════════════════════════

ACCORDS_HEADERS = {
    'ACCORDS': 2, 'PARAMETRES': 0, 'PARTENAIRES': 0,
    'SUIVI_ACCORDS': 2, 'DECAISSEMENTS': 2, 'AVENANTS': 1, 'COMMISSIONS_MIXTES': 1,
}

def read_accords_sheet(sheet_name):
    if not os.path.exists(ACCORDS_FILE):
        return None
    hdr = ACCORDS_HEADERS.get(sheet_name, 0)
    return pd.read_excel(ACCORDS_FILE, sheet_name=sheet_name, dtype=str, header=hdr)

@app.route('/api/partenaires/list', methods=['GET'])
@require_auth
def get_partenaires():
    try:
        df = read_accords_sheet('PARTENAIRES')
        if df is None: return jsonify([])
        df.columns = [str(c).strip() for c in df.columns]
        df = df.replace({np.nan: None})
        return jsonify(json.loads(df.to_json(orient='records', force_ascii=False)))
    except Exception as e:
        return jsonify({"error": _safe_error(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# MODULE COOPÉRATION DÉCENTRALISÉE - NON MIGRE
# ══════════════════════════════════════════════════════════════════════════════

COOP_DEC_FILE = os.path.join(BASE_DIR, 'Cooperation_Decentralisee.xlsx')

def init_coop_dec_excel():
    if not os.path.exists(COOP_DEC_FILE):
        with pd.ExcelWriter(COOP_DEC_FILE, engine='openpyxl') as writer:
            headers_accords = ['Ref_Accord', 'Commune_Benin', 'Departement_Benin',
                'Collectivite_Partenaire', 'Pays_Partenaire', 'Type_Cooperation',
                'Date_Signature', 'Date_Debut', 'Date_Fin', 'Duree_Annees',
                'Secteurs_Intervention', 'Objectifs', 'Montant_Total', 'Devise',
                'Contributions_Benin', 'Contributions_Partenaire', 'Statut',
                'Point_Focal_Benin', 'Contact_Partenaire', 'Frequence_Reunions',
                'Derniere_Reunion', 'Prochaine_Reunion', 'Actions_Realisees',
                'Echanges_Effectues', 'Resultats_Obtenus', 'Difficultes',
                'Recommandations', 'Fichier_Convention', 'Observations']
            pd.DataFrame(columns=headers_accords).to_excel(writer, sheet_name='ACCORDS', index=False)
            headers_actions = ['Ref_Action', 'Ref_Accord', 'Commune_Benin', 'Titre_Action',
                'Secteur', 'Date_Debut', 'Date_Fin', 'Budget', 'Financeur_Principal',
                'Beneficiaires', 'Resultats', 'Statut']
            pd.DataFrame(columns=headers_actions).to_excel(writer, sheet_name='ACTIONS_PROJETS', index=False)
            headers_echanges = ['Ref_Echange', 'Ref_Accord', 'Type_Echange', 'Date',
                'Lieu', 'Participants_Benin', 'Participants_Partenaires', 'Objectifs', 'Resultats', 'Rapport']
            pd.DataFrame(columns=headers_echanges).to_excel(writer, sheet_name='ECHANGES_VISITES', index=False)
            params_data = {
                'TYPE_COOPERATION': ['Jumelage', 'Partenariat', "Accord d'amitié", 'Coopération triangulaire'],
                'STATUT_COOP': ['Actif', 'Expiré', 'Suspendu', 'En négociation'],
                'SECTEURS': ['Eau & Assainissement', 'Éducation', 'Santé', 'Agriculture', 'Gouvernance', 'Culture', 'Environnement'],
                'COMMUNES_BENIN': ['Cotonou', 'Porto-Novo', 'Parakou', 'Bohicon', 'Abomey-Calavi', 'Djougou'],
                'PAYS': ['France', 'Belgique', 'Canada', 'Allemagne', 'Sénégal', "Côte d'Ivoire"]
            }
            pd.DataFrame(dict([(k, pd.Series(v)) for k, v in params_data.items()])).to_excel(writer, sheet_name='PARAMETRES', index=False)

init_coop_dec_excel()

def read_coop_sheet(sheet_name):
    if not os.path.exists(COOP_DEC_FILE): return None
    return pd.read_excel(COOP_DEC_FILE, sheet_name=sheet_name, dtype=str, header=0)

@app.route('/api/coop_dec/list', methods=['GET'])
@require_auth
def get_coop_dec_list():
    try:
        df = read_coop_sheet('ACCORDS')
        if df is None or df.empty: return jsonify([])
        df = df.replace({np.nan: None})
        for col in ['Date_Signature', 'Date_Debut', 'Date_Fin', 'Derniere_Reunion', 'Prochaine_Reunion']:
            if col in df.columns: df[col] = df[col].apply(clean_date)
        return jsonify(json.loads(df.to_json(orient='records', force_ascii=False)))
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500

@app.route('/api/coop_dec/parametres', methods=['GET'])
@require_auth
def get_coop_dec_parametres():
    try:
        df = read_coop_sheet('PARAMETRES')
        if df is None: return jsonify({})
        params = {col: [str(v).strip() for v in df[col].dropna() if str(v).strip()] for col in df.columns}
        return jsonify(params)
    except Exception as e:
        return jsonify({"error": _safe_error(e)}), 500

@app.route('/api/coop_dec/actions', methods=['GET'])
@require_auth
def get_coop_dec_actions():
    try:
        df = read_coop_sheet('ACTIONS_PROJETS')
        if df is None: return jsonify([])
        df = df.replace({np.nan: None})
        return jsonify(json.loads(df.to_json(orient='records', force_ascii=False)))
    except Exception as e:
        return jsonify({"error": _safe_error(e)}), 500

@app.route('/api/coop_dec/echanges', methods=['GET'])
@require_auth
def get_coop_dec_echanges():
    try:
        df = read_coop_sheet('ECHANGES_VISITES')
        if df is None: return jsonify([])
        df = df.replace({np.nan: None})
        return jsonify(json.loads(df.to_json(orient='records', force_ascii=False)))
    except Exception as e:
        return jsonify({"error": _safe_error(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# AUTO-MIGRATION: Création des tables si elles n'existent pas
# ══════════════════════════════════════════════════════════════════════════════

def _get_pg_connection():
    """Connexion directe PostgreSQL via psycopg2 pour DDL."""
    import psycopg2
    url = os.environ.get('SUPABASE_URL', '')
    # Extraire le project_ref de l'URL Supabase
    # https://xxxxx.supabase.co -> xxxxx
    project_ref = url.replace('https://', '').replace('.supabase.co', '').split('/')[0]
    db_password = os.environ.get('SUPABASE_DB_PASSWORD', '')
    if not db_password:
        raise RuntimeError("SUPABASE_DB_PASSWORD manquant dans l'environnement")
    return psycopg2.connect(
        host=f'db.{project_ref}.supabase.co',
        port=5432,
        dbname='postgres',
        user='postgres',
        password=db_password
    )


@app.route('/api/setup/create-tables', methods=['POST'])
@require_auth
def setup_create_tables():
    """Crée les tables manquantes dans Supabase (one-time setup)."""
    try:
        conn = _get_pg_connection()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS secteur_sous_secteur (
                id              SERIAL PRIMARY KEY,
                secteur         TEXT NOT NULL,
                sous_secteur    TEXT NOT NULL,
                created_at      TIMESTAMPTZ DEFAULT NOW(),
                UNIQUE(secteur, sous_secteur)
            );
            CREATE INDEX IF NOT EXISTS idx_secteur_sous ON secteur_sous_secteur(secteur);
        """)
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Tables créées avec succès"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": _safe_error(e)}), 500


@app.route('/api/banque/parametres/diag', methods=['GET'])
@require_auth
def diag_parametres():
    """Diagnostic: verifie l'etat de la table parametres."""
    result = {"status": "unknown", "details": {}}
    try:
        from db import get_supabase
        sb = get_supabase()
        # Test SELECT
        resp = sb.table('parametres').select('id', 'categorie', 'valeur').limit(5).execute()
        result["details"]["select"] = "OK"
        result["details"]["sample"] = resp.data
        # Test INSERT (puis rollback)
        test_val = f"__diag_test_{int(time.time())}"
        try:
            sb.table('parametres').insert({'categorie': '__diag', 'valeur': test_val, 'ordre': 0}).execute()
            result["details"]["insert"] = "OK"
            # Nettoyer
            sb.table('parametres').delete().eq('categorie', '__diag').eq('valeur', test_val).execute()
            result["details"]["delete"] = "OK"
        except Exception as e2:
            result["details"]["insert"] = f"FAIL: {str(e2)}"
            result["status"] = "read_only"
        if result["status"] != "read_only":
            result["status"] = "ok"
    except Exception as e:
        result["status"] = "error"
        result["details"]["error"] = str(e)
    return jsonify(result)


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import socket
    local_ip = socket.gethostbyname(socket.gethostname())
    print("=" * 60)
    print("  SGIAD - Serveur de la Plateforme Nationale")
    print("=" * 60)
    print(f"  Acces local   : http://127.0.0.1:5000")
    print(f"  Acces reseau  : http://{local_ip}:5000")
    print(f"  Partagez cette URL avec vos collegues :")
    print(f"  >>> http://{local_ip}:5000 <<<")
    print("=" * 60)
    app.run(host='0.0.0.0', debug=os.environ.get('FLASK_DEBUG') == '1', port=5000)
