# -*- coding: utf-8 -*-
"""
Module API — GESTION DES ACCORDS FINANCIERS (v2 Supabase)
==========================================================
Remplace l'ancien module Excel par Supabase (PostgreSQL cloud).
Toutes les donnees viennent de la table accords_consolides.
La metadata des colonnes est stockee dans colonnes_meta.

Routes API :
    GET  /api/accords-financiers/columns
    GET  /api/accords-financiers/list
    POST /api/accords-financiers/add
    PUT  /api/accords-financiers/update/<code>
    DEL  /api/accords-financiers/delete/<code>
    POST /api/accords-financiers/export
    POST /api/accords-financiers/import
"""

import os
import json
import traceback
from io import BytesIO
from datetime import datetime

from flask import Blueprint, request, jsonify, send_file

from auth import require_auth

bp = Blueprint('accords_financiers', __name__)

KEY_HEADER = 'Code Projet'


def _safe_error(e):
    """Message d'erreur sur : details techniques seulement en debug local."""
    if os.environ.get('FLASK_DEBUG') == '1':
        return f"{type(e).__name__}: {str(e)}"
    return "Une erreur interne est survenue. Veuillez réessayer."


# ══════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════

@bp.route('/api/accords-financiers/columns', methods=['GET'])
def get_columns():
    """Retourne la metadata des colonnes pour le frontend."""
    try:
        from db import get_columns_meta
        columns = get_columns_meta()
        return jsonify({'key_field': KEY_HEADER, 'columns': columns})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


@bp.route('/api/accords-financiers/list', methods=['GET'])
@require_auth
def list_accords():
    """Retourne tous les accords."""
    try:
        from db import accords_list
        rows = accords_list()
        return jsonify({'total': len(rows), 'data': rows})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


@bp.route('/api/accords-financiers/add', methods=['POST'])
@require_auth
def add_accord():
    """Ajoute un nouvel accord."""
    try:
        from db import accords_insert, accords_get
        payload = request.get_json(force=True) or {}

        code = payload.get(KEY_HEADER)
        if not code or str(code).strip() == '':
            return jsonify({'error': f'"{KEY_HEADER}" est obligatoire.'}), 400

        # Verifier si le code existe deja
        existing = accords_get(str(code).strip())
        if existing:
            return jsonify({'error': f'Un accord avec le code "{code}" existe déjà.'}), 409

        result = accords_insert(payload)
        return jsonify({'message': 'Accord enregistré.', KEY_HEADER: code}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


@bp.route('/api/accords-financiers/update/<code>', methods=['PUT'])
@require_auth
def update_accord(code):
    """Met a jour un accord existant."""
    try:
        from db import accords_update, accords_get
        from urllib.parse import unquote
        code = unquote(code).strip()
        payload = request.get_json(force=True) or {}

        # Verifier que l'accord existe
        existing = accords_get(code)
        if not existing:
            return jsonify({'error': f'Accord "{code}" introuvable.'}), 404

        # Si le code change, verifier que le nouveau n'existe pas
        new_code = payload.get(KEY_HEADER, code)
        if str(new_code).strip() != code:
            conflict = accords_get(str(new_code).strip())
            if conflict:
                return jsonify({'error': f'Le code "{new_code}" est déjà utilisé.'}), 409

        accords_update(code, payload)
        return jsonify({'message': 'Accord mis à jour.'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


@bp.route('/api/accords-financiers/delete/<code>', methods=['DELETE'])
@require_auth
def delete_accord(code):
    """Supprime un accord."""
    try:
        from db import accords_delete, accords_get
        from urllib.parse import unquote
        code = unquote(code).strip()

        existing = accords_get(code)
        if not existing:
            return jsonify({'error': f'Accord "{code}" introuvable.'}), 404

        accords_delete(code)
        return jsonify({'message': 'Accord supprimé.'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


@bp.route('/api/accords-financiers/export', methods=['POST'])
@require_auth
def export_accords():
    """Exporte les accords filtres en Excel."""
    try:
        from db import accords_list, get_columns_meta
        payload = request.get_json(force=True) or {}
        filters = payload.get('filters', {})
        selected_columns = payload.get('selected_columns', [])

        all_rows = accords_list()
        columns_meta = get_columns_meta()
        columns_by_key = {c['key']: c for c in columns_meta}

        # Filtrer les lignes
        def matches(row):
            for key, allowed in filters.items():
                if not allowed:
                    continue
                allowed_set = {str(v).strip() for v in allowed if str(v).strip()}
                if not allowed_set:
                    continue
                if str(row.get(key, '')).strip() not in allowed_set:
                    return False
            return True

        filtered = [r for r in all_rows if matches(r)]
        cols_to_export = selected_columns or list(columns_by_key.keys())

        # Generer le fichier Excel
        from openpyxl import Workbook
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = 'Accords_Export'
        out_ws.append(cols_to_export)
        for r in filtered:
            out_ws.append([r.get(col, '') for col in cols_to_export])

        buf = BytesIO()
        out_wb.save(buf)
        buf.seek(0)
        filename = f"Export_Accords_Financiers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500


@bp.route('/api/accords-financiers/import', methods=['POST'])
@require_auth
def import_accords():
    """Importe des accords depuis un fichier Excel."""
    try:
        from db import accords_insert, accords_update, accords_get
        from openpyxl import load_workbook

        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'Aucun fichier reçu.'}), 400

        in_wb = load_workbook(BytesIO(file.read()), data_only=True)
        in_ws = in_wb.active

        # Lire les en-tetes
        in_headers = []
        col_idx = 1
        while True:
            h = in_ws.cell(row=1, column=col_idx).value
            if h is None or str(h).strip() == '':
                break
            in_headers.append(str(h).strip())
            col_idx += 1

        if KEY_HEADER not in in_headers:
            return jsonify({'error': f'Le fichier doit contenir "{KEY_HEADER}".'}), 400

        added, updated = 0, 0

        for r in range(2, in_ws.max_row + 1):
            row_data = {}
            empty = True
            for i, h in enumerate(in_headers, start=1):
                v = in_ws.cell(row=r, column=i).value
                if v is not None and str(v).strip() != '':
                    empty = False
                row_data[h] = v
            if empty:
                continue

            code = row_data.get(KEY_HEADER)
            if not code:
                continue

            existing = accords_get(str(code).strip())
            if existing:
                accords_update(str(code).strip(), row_data)
                updated += 1
            else:
                accords_insert(row_data)
                added += 1

        return jsonify({'message': f'{added} accord(s) ajouté(s), {updated} mis à jour.'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': _safe_error(e)}), 500
